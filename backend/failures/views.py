import requests
import re
import html
import json
from io import BytesIO
import openpyxl

from django.conf import settings
from django.utils import timezone # <-- IMPORT IS NEEDED
from django.http import HttpResponse

from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.views import APIView

from notifications.models import TelegramGroup # <-- IMPORT IS NEEDED
from .models import Failure, FailureAttachment
from .serializers import FailureListSerializer, FailureCreateUpdateSerializer, FailureAttachmentSerializer

from openpyxl.styles import PatternFill, Alignment, Font
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape


class FailureViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows failures to be viewed or edited.
    """
    queryset = Failure.objects.filter(is_archived=False).select_related(
        'circuit', 'station', 'section', 'sub_section', 'assigned_to'
    ).all()
    
    permission_classes = [permissions.AllowAny]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return FailureCreateUpdateSerializer
        return FailureListSerializer

    @action(detail=True, methods=['post'])
    def archive(self, request, pk=None):
        failure = self.get_object()
        failure.is_archived = True
        failure.archived_at = timezone.now()
        failure.archived_reason = request.data.get('reason', '')
        failure.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'])
    def notify(self, request, pk=None):
        failure = self.get_object()
        group_keys = request.data.get('groups', [])
        if not group_keys:
            return Response({"error": "No notification groups provided."}, status=status.HTTP_400_BAD_REQUEST)
        # ... (rest of notify method is unchanged) ...
        status_text = failure.current_status
        status_emoji = ''
        if status_text == 'Active': status_emoji = ' 🔴'
        elif status_text == 'Resolved': status_emoji = ' ✅'
        elif status_text == 'In Progress': status_emoji = ' 🟡'
        elif status_text == 'On Hold': status_emoji = ' 🟠'
        elif status_text == 'Information': status_emoji = ' ℹ️'
        message_parts = [f"*ID:* {failure.fail_id}", "", f"*Circuit:* {failure.circuit.circuit_id if failure.circuit else 'N/A'}", f"*Status:* {status_text}{status_emoji}", ""]
        station_display = (failure.station.code or failure.station.name) if failure.station else 'N/A'
        section_display = failure.section.name if failure.section else 'N/A'
        sub_section_display = failure.sub_section.name if failure.sub_section else 'N/A'
        if failure.entry_type == 'item' or (station_display not in ['ALL', 'N/A']):
            message_parts.append(f"*Station:* {station_display}")
        if failure.entry_type == 'item' or (section_display not in ['ALL', 'N/A']):
            message_parts.append(f"*Section:* {section_display}")
        if failure.sub_section and (failure.entry_type == 'item' or (sub_section_display not in ['ALL', 'N/A'])):
            message_parts.append(f"*Sub-Section:* {sub_section_display}")
        message_parts.append("")
        if failure.remark_fail:
            fail_remarks = (failure.remark_fail[:500] + '...') if len(failure.remark_fail) > 500 else failure.remark_fail
            message_parts.append(f"❗️ *Fail Remarks:*\n{fail_remarks}")
        if failure.current_status == 'Resolved' and failure.remark_right:
            resolved_remarks = (failure.remark_right[:500] + '...') if len(failure.remark_right) > 500 else failure.remark_right
            if failure.remark_fail:
                message_parts.append("")
            message_parts.append(f"✅ *Resolved Remark:*\n{resolved_remarks}")
        message = "\n".join(message_parts)
        try:
            groups_to_notify = TelegramGroup.objects.filter(key__in=group_keys)
            for group in groups_to_notify:
                group.send_message(message)
            return Response({"message": "Notifications sent successfully."})
        except Exception as e:
            return Response({"error": f"Failed to send notification: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        failure_ids = request.query_params.get('ids', None)
        if failure_ids:
            ids = [int(id) for id in failure_ids.split(',')]
            queryset = self.get_queryset().filter(id__in=ids)
        else:
            queryset = self.get_queryset()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Failure Logs'
        sheet.merge_cells('A1:I1')
        title_cell = sheet['A1']
        title_cell.value = 'Recent Failure Report'
        title_cell.font = Font(size=18, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        sheet.merge_cells('A2:I2')
        timestamp_cell = sheet['A2']
        timestamp_cell.value = f"Generated on: {timezone.now().strftime('%d-%b-%Y %H:%M')}"
        timestamp_cell.font = Font(italic=True)
        timestamp_cell.alignment = Alignment(horizontal='center')
        sheet.append([])
        headers = ["Fail ID", "Circuit", "Station", "Section", "Sub Section", "Reported At", "Resolved At", "Duration", "Remarks"]
        sheet.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in sheet[4]:
            cell.font = header_font
            cell.fill = header_fill
        status_colors = {'Active': 'FFC7CE', 'In Progress': 'FFEB9C', 'Resolved': 'C6EFCE', 'On Hold': 'FFD9A5'}
        thin_border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
        for failure in queryset:
            duration_str = '—'
            if failure.reported_at and failure.resolved_at:
                duration_delta = failure.resolved_at - failure.reported_at
                hours, remainder = divmod(duration_delta.total_seconds(), 3600)
                minutes, _ = divmod(remainder, 60)
                duration_str = f"{int(hours)}h {int(minutes)}m"
            station_code = failure.station.code if failure.station and failure.station.code else (failure.station.name if failure.station else 'N/A')
            row_data = [
                failure.fail_id, failure.circuit.circuit_id if failure.circuit else 'N/A',
                station_code, failure.section.name if failure.section else 'N/A',
                failure.sub_section.name if failure.sub_section else 'N/A',
                failure.reported_at.strftime("%Y-%m-%d %H:%M") if failure.reported_at else '',
                failure.resolved_at.strftime("%Y-%m-%d %H:%M") if failure.resolved_at else '',
                duration_str, failure.remark_fail
            ]
            sheet.append(row_data)
            fill_color = status_colors.get(failure.current_status)
            for cell in sheet[sheet.max_row]:
                cell.border = thin_border
                if fill_color:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        sheet.column_dimensions['A'].width = 15; sheet.column_dimensions['B'].width = 15; sheet.column_dimensions['C'].width = 12; sheet.column_dimensions['D'].width = 20; sheet.column_dimensions['E'].width = 20; sheet.column_dimensions['F'].width = 18; sheet.column_dimensions['G'].width = 18; sheet.column_dimensions['H'].width = 12; sheet.column_dimensions['I'].width = 50
        for row in sheet.iter_rows(min_row=5):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        buffer = BytesIO()
        workbook.save(buffer)
        try:
            reports_group = TelegramGroup.objects.get(key='reports')
            filename = f"failure_report_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
            caption = f"Failure log report (Excel) generated on {timezone.now().strftime('%d-%b-%Y %H:%M')}."
            buffer.seek(0)
            reports_group.send_document(file_object=buffer, filename=filename, caption=caption)
        except TelegramGroup.DoesNotExist:
            print("Warning: 'reports' Telegram group not found. Skipping Telegram send.")
        except Exception as e:
            print(f"Error sending Excel report to Telegram: {e}")
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=failure_logs_{timezone.now().strftime("%Y%m%d")}.xlsx'
        return response

    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        failure_ids = request.query_params.get('ids', None)
        if failure_ids:
            ids = [int(id) for id in failure_ids.split(',')]
            queryset = self.get_queryset().filter(id__in=ids)
        else:
            queryset = self.get_queryset()

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        body_style = styles['BodyText']
        styles['h1'].alignment = 1
        elements = []

        title = "Recent Failure Report"
        elements.append(Paragraph(title, styles['h1']))
        timestamp = f"Generated on: {timezone.now().strftime('%d-%b-%Y %H:%M')}"
        elements.append(Paragraph(timestamp, styles['Normal']))
        elements.append(Spacer(1, 24))

        status_colors_pdf = {'Active': colors.lightpink, 'In Progress': colors.lightyellow, 'Resolved': colors.lightgreen, 'On Hold': colors.peachpuff}

        table_style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
        ]

        data = [['Fail ID', 'Circuit', 'Station', 'Section', 'Sub Section', 'Reported', 'Resolved', 'Duration', 'Remarks']]
        for i, failure in enumerate(queryset):
            color = status_colors_pdf.get(failure.current_status)
            if color:
                table_style_commands.append(('BACKGROUND', (0, i + 1), (-1, i + 1), color))

            station_code = failure.station.code if failure.station and failure.station.code else (failure.station.name if failure.station else 'N/A')
            duration_str = '—'
            if failure.reported_at and failure.resolved_at:
                duration_delta = failure.resolved_at - failure.reported_at
                hours, remainder = divmod(duration_delta.total_seconds(), 3600)
                minutes, _ = divmod(remainder, 60)
                duration_str = f"{int(hours)}h {int(minutes)}m"

            data.append([
                failure.fail_id,
                failure.circuit.circuit_id if failure.circuit else 'N/A',
                station_code,
                Paragraph(failure.section.name if failure.section else 'N/A', body_style),
                Paragraph(failure.sub_section.name if failure.sub_section else 'N/A', body_style),
                failure.reported_at.strftime("%d-%b-%y %H:%M") if failure.reported_at else '',
                failure.resolved_at.strftime("%d-%b-%y %H:%M") if failure.resolved_at else '',
                duration_str,
                Paragraph(failure.remark_fail, body_style),
            ])

        table = Table(data, colWidths=[80, 60, 50, 80, 80, 90, 90, 60, 150])
        table.setStyle(TableStyle(table_style_commands))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        try:
            reports_group = TelegramGroup.objects.get(key='reports')
            filename = f"failure_report_{timezone.now().strftime('%Y-%m-%d')}.pdf"
            caption = f"Failure log report (PDF) generated on {timezone.now().strftime('%d-%b-%Y %H:%M')}."
            buffer.seek(0)
            reports_group.send_document(file_object=buffer, filename=filename, caption=caption)
        except TelegramGroup.DoesNotExist:
            print("Warning: 'reports' Telegram group not found. Skipping Telegram send for PDF.")
        except Exception as e:
            print(f"Error sending PDF report to Telegram: {e}")

        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="failure_logs_{timezone.now().strftime("%Y%m%d")}.pdf"'
        return response

class FailureAttachmentTelegramView(APIView):
    # ... (content remains the same)
    pass

class FailureAttachmentViewSet(viewsets.ModelViewSet):
    queryset = FailureAttachment.objects.all()
    serializer_class = FailureAttachmentSerializer
    parser_classes = [MultiPartParser, FormParser]
    filterset_fields = ['failure']

class ArchivedFailureViewSet(viewsets.ModelViewSet):
    serializer_class = FailureListSerializer
    queryset = Failure.objects.filter(is_archived=True).select_related(
        'circuit', 'station', 'section', 'assigned_to'
    ).order_by('-archived_at')
    permission_classes = [permissions.AllowAny]

class SendAttachmentToTelegramView(APIView):
    """
    Accepts a file upload and immediately sends it to specified Telegram groups
    without saving the attachment to the database.
    """
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [permissions.AllowAny] # Adjust in production

    def post(self, request, *args, **kwargs):
        failure_id = request.data.get('failure_id')
        file_obj = request.FILES.get('file')
        # The group_keys might be a JSON string from FormData
        group_keys_str = request.data.get('group_keys', '[]')

        if not all([failure_id, file_obj]):
            return Response(
                {"error": "A failure_id and a file are required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            group_keys = json.loads(group_keys_str)
            if not isinstance(group_keys, list):
                raise ValueError("group_keys must be a list.")
        except (json.JSONDecodeError, ValueError) as e:
            return Response(
                {"error": f"Invalid format for group_keys: {e}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            failure = Failure.objects.get(pk=failure_id)
            target_groups = list(TelegramGroup.objects.filter(key__in=group_keys))
            if not target_groups:
                return Response({"error": "No valid Telegram groups found."}, status=status.HTTP_400_BAD_REQUEST)

            # --- Construct the caption ---
            circuit_id_str = failure.circuit.circuit_id if failure.circuit else 'N/A'
            station_display_str = failure.station.code if failure.station and failure.station.code else (failure.station.name if failure.station else 'N/A')
            duration_str = 'N/A'
            if failure.reported_at and failure.resolved_at:
                duration_delta = failure.resolved_at - failure.reported_at
                hours, remainder = divmod(duration_delta.total_seconds(), 3600)
                minutes, _ = divmod(remainder, 60)
                duration_str = f"{int(hours)}h {int(minutes)}m"

            def sanitize_for_tag(text):
                return re.sub(r'[^a-zA-Z0-9_]', '', str(text).replace(' ', '_')) if text else ''

            caption_parts = [
                f"<b>File Uploaded for Failure ID:</b> {failure.fail_id}",
                f"<b>Circuit:</b> {circuit_id_str}",
                f"<b>Station:</b> {station_display_str}",
                f"<b>Resolution Remarks:</b> {failure.remark_right}" if failure.remark_right else None,
                f"<b>Duration:</b> {duration_str}",
                "---",
                f"#{sanitize_for_tag(circuit_id_str)} #{sanitize_for_tag(station_display_str)}"
            ]
            caption = "\n".join(filter(None, caption_parts))
            # --- End caption construction ---

            for group in target_groups:
                file_obj.seek(0)
                group.send_document(
                    file_object=file_obj,
                    filename=file_obj.name,
                    caption=caption
                )
            return Response({"message": f"File sent to {len(target_groups)} group(s)."}, status=status.HTTP_200_OK)

        except Failure.DoesNotExist:
            return Response({"error": "Failure not found."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": f"Failed to send document: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
